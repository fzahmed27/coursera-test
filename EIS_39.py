import tkinter as tk
from tkinter import ttk
import openpyxl
from PIL import Image, ImageTk



 
"""GUI for Enigneering info sheet which will store data in
database. Entries in the filed are written to variable"""
 
 
 
LARGE_FONT = ("Helvetica",12, "bold")
MIDDLE_FONT = ("Helvetica", 10, "bold")
NORM_FONT = ("Helvetica", 10)
row1 = 1
mat_prop = {1:110000, 2: 85000}
input_stage1 = []
input_stage2 = []
input_stage3 = []
input_stage4=  []
input_col_2 = []
input_col_3 = []
input_col_4 = []
input_col_5 = []
valve_dim_stage1 = []
valve_dim_stage2 = []
valve_dim_stage3 = []
valve_dim_stage4 = []
state = 0
stage_suc ={ "STAGE 1" :"1S", "STAGE 2":"2S", "STAGE 3": "3S", "STAGE 4": "4S"}
stage_dis ={ "STAGE 1" :"1D", "STAGE 2":"2D", "STAGE 3": "3D", "STAGE 4": "4D"}
pocket_depth =[0, 0, 0]
pocket_depth2 =[0, 0, 0]
comp_data =[]
seat_thick = 0
carrier_thick = 0 
valve_thick = 0
stage = "STAGE 1"
modules_suc = 0
modules_dis = 0
stage_data_col2 =[]
isedit = False



def popupmsg(msg):

    #This is  popupmsg which could be used to inform the user of any event
    popup = tk.Tk()
    popup.wm_title("!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.pack(side="top", fill="x", pady=10)
    B1 = ttk.Button(popup, text="Okay", command = popup.destroy)
    B1.pack()
    popup.mainloop()
    


class InfoSheet(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        container = tk.Frame(self)

        container.pack(side="top", fill="both", expand = True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        menubar = tk.Menu(container)
        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Save settings", command = lambda: popupmsg("Not supported just yet!"))
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command= quit)
        menubar.add_cascade(label="File", menu=filemenu)
    
        tk.Tk.config(self, menu=menubar)
    
        self.frames = {}
        # shows pages 
        for F in (StartPage, PageTwo):

            frame = F(container, self)

            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(StartPage)
  
    def show_frame(self, cont):

        frame = self.frames[cont]
        frame.tkraise()
        

#create startpage for the application. This is the first page when application is executed      
class StartPage(tk.Frame):
    def __init__(self, parent, controller):

        #----This is the first window which inherits from parent window ----
        tk.Frame.__init__(self, parent)
        self.canvas = tk.Canvas(self, borderwidth=0, background="lightgrey")
        self.frame  = tk.Frame(self.canvas, background="#CCF1FF", borderwidth =3, relief = "groove")
        self.frame1 = tk.Frame(self.canvas, background="#CCF1FF", borderwidth =3, relief = "groove")
        self.frame2 = tk.Frame(self.canvas, background="#CCF1FF", borderwidth =3, relief = "groove")
        self.frame3 = tk.Frame(self.canvas, background="#CCF1FF", borderwidth =3, relief = "groove")
        self.frame4 = tk.Frame(self.canvas, background="#CCF1FF", borderwidth =3, relief = "groove")
        self.frame5 = tk.Frame(self.canvas, background="#CCF1FF", borderwidth =3, relief = "groove")
        
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand= True)
        self.canvas.create_window((170,55), window=self.frame, anchor="nw", 
                                  tags="self.frame")

        #Frame 1 consist of user data 
        self.canvas.create_window((150,560),window = self.frame3, anchor ="nw", tags = "self.frame1")
        #
        self.canvas.create_window((170,220),window = self.frame1, anchor ="nw", tags = "self.frame1")
        self.canvas.create_window((275,480),window = self.frame2, anchor ="nw", tags = "self.frame3")
        self.canvas.create_window((90,780),window = self.frame4, anchor ="nw", tags = "self.frame4")
        self.canvas.create_window((90,1000),window = self.frame5, anchor ="nw", tags = "self.frame5")

        self.frame.bind("<Configure>", self.onFrameConfigure)
        
        #creating a button for Page and save the data to excel sheet
        button1 = ttk.Button(self.canvas, text = "SAVE", command =self.get_edit_data)
        
        self.canvas.create_window((770, 1100), window = button1, anchor = "nw")
        button2 = ttk.Button(self.canvas, text = "NEXT", command=lambda: controller.show_frame(PageOne))
        self.canvas.create_window((850, 1100), window = button2, anchor = "nw")                                                                       
         

         # creating labels for the entry field
        tk.Label(self.frame, text = "COMPRESSOR DATA", font = MIDDLE_FONT, fg ="blue", bg ="white").grid(row =5,
                                                                                column =2, sticky ="w")
        
        tk.Label(self.frame1, text = "OPERATING DATA", font = MIDDLE_FONT, fg ="blue", bg ="white").grid(row =1,
                                                                                                         column =2, sticky ="w")
 
        
        
        #label for vale design data
        #label13 = tk.Label(self.frame3, text = "VALVE DESIGN DATA ", font = MIDDLE_FONT, fg ="blue", bg = "white")
        #label13.grid(row =26, column =2, sticky ="we")
        #label15 =  tk.Label(self.frame1, text = "STAGE 1", font = NORM_FONT, bg = "grey", width = 10)
        #label15.grid(row =3, column =1,padx=2,  pady=2,  sticky ="we")
        #label14= tk.Label(self.frame2, text = "MATERIAL DATA ",font = MIDDLE_FONT, fg = "blue", bg ="white")
        #label14.grid(row =21, column =1, sticky ="w")
        label_frame1 = tk.Label(self.canvas,text = "MATERIAL DATA", fg = "blue", font = MIDDLE_FONT, bg = "white")
        self.canvas.create_window(410,460, window = label_frame1, anchor = "nw")      
        label_frame2 = tk.Label(self.canvas,text = "SEAT DESIGN DATA ", fg = "blue", font = MIDDLE_FONT, bg = "white")
        self.canvas.create_window(410,535, window = label_frame2, anchor = "nw")
        label_frame3 = tk.Label(self.canvas,text = "VALVE DESIGN DATA ", fg = "blue", font = MIDDLE_FONT, bg = "white")
        self.canvas.create_window(410,753, window = label_frame3, anchor = "nw")
        label_frame4 = tk.Label(self.canvas,text = "ZVI VALVE DIM ", fg = "blue", font = MIDDLE_FONT, bg = "white")
        self.canvas.create_window(410,975, window = label_frame4, anchor = "nw")



        self.variables=[]
        #----Global varaibles for storing the data from entry fields for base case spare data structure if needed
        self.field_stage4 = []
        self.field_stage3 = []
        self.field_stage2 = []
        self.field_stage1 = []
        self.field_stage5 = []
        self.field_stage6 = []
		#-----data structure for getting values when the values are read from excel and edited
        self.edit_data_label3 = [] 
        self.edit_data_label5 = []
        self.edit_data_label8 = []
        self.edit_data_label9 = []
        
        #----Global varaibles for storing the data from entry fields for operting condition 
        self.field_stage4_oper = []
        self.field_stage3_oper = []
        self.field_stage2_oper = []
        self.field_stage1_oper = []
        self.field_stage5_oper = []
        self.field_stage6_oper = []
        #----Global varaibles for storing the data from entry fields for pocket_depth
        self.field_stage4_pd = []
        self.field_stage3_pd = []
        self.field_stage2_pd = []
        self.field_stage1_pd = []
        self.field_stage5_pd = []
        self.field_stage6_pd = []
        #----Global varaibles for storing the data from entry fields for seat_desin objects
        self.field_stage4_sd = []
        self.field_stage3_sd = []
        self.field_stage2_sd = []
        self.field_stage1_sd = []
        self.field_stage5_sd = []
        self.field_stage6_sd = []
    #----Global varaibles for storing the data from entry fields for valve thickness parameter
        self.field_stage4_vd = []
        self.field_stage3_vd = []
        self.field_stage2_vd = []
        self.field_stage1_vd = []
        self.field_stage5_vd = []
        self.field_stage6_vd = []
        # all labels are stored in the following list 
        self.label_1 = []
        self.label_2 = []
        self.label_3 = []
        self.label_4 = []
        self.label_5  =[]

        #Row values for the labels and entry fields associated with it
        self.rows1 = [4, 5, 6, 7, 8, 9, 10]
        self.rows2 = [3,4,5,6,7,8]
        self.rows3 = [2,3,4,5,6,7,8,9]
        self.rows4 = [2,3,4]

        self.create_label1()
        self.create_label2()
        self.create_label3()
        self.create_label4()
        self.create_label5()
        self.create_label7()
        self.create_label8()
        self.create_label9()
        self.create_label10()
        
        self.fos()
        self.showImg_logo()
        self.choose_stage()
        self.delete_button()
     
        
    def onFrameConfigure(self, event):
        #Reset the scroll region to encompass the inner frame
        self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        
    def showImg_logo(self):
        self.image2 = tk.PhotoImage(file = "company_logo.png")
        self.canvas.create_image(50,0, image= self.image2, anchor ="nw")

    def choose_stage(self):
        
        var1 = tk.IntVar()
        choices = [ 2, 2, 3, 4, 5, 6]
        drop1 = ttk.OptionMenu(self.frame1,var1,*choices)
        drop1["menu"].config(bg = "white")
        drop1.grid(row = 2, column = 2, sticky = "w")
        stage = "STAGE 1"
        
        print ("column choosen" , var1.get())
        button14 =  ttk.Button(self.frame1, text = "ENTER STAGE", width = 15)
        button14.grid(row =2, column =1, sticky ="e")
        button14.bind("<Button-1>", lambda event:self.display_stage_frame1(var1.get()))
        button14.bind("<Button-1>", lambda event:self.display_stage_frame2(stage, var1.get()), add="+")
        
        
    def on_button(self):
        """ This function saves all the data fro entry filed into list """        
        while True:
                try:
                        sq_num = int(self.record[0].get())
                        wb = wb = openpyxl.load_workbook(filename='example1.xlsx', data_only=True)
                        ws1 = wb.get_sheet_by_name('Sheet1')
            
                        row1 =1
                        for row in ws1.iter_rows(row_offset = 1):
                            row1 +=1
                            if sq_num == row[0].value:
                               return self.read_data(row1)
                            elif row[0].value is None:
                                  return self.get_data(row1, sq_num)
                except ValueError:
                        popupmsg("Enter Serial No")
        
    
       
    def create_label1(self):
        row = [3,4]
        field1 = ["PREPARED BY", "DATE"]
        row = 1
        self.values1 = []
        for item in field1:
            i = len(self.values1)
            label = tk.Label(self.frame, text = item, bg = "grey", font=("Helvetica", 8), borderwidth="1")
            label.grid(row = row, column = 0 , sticky ="e")
            self.values1.append(tk.StringVar())
            self.values1[i].trace("w", lambda name, index, mode, var=self.values1[i], i=i:self.get_comp_data(var, i)) 
            entry1 =tk.Entry(self.frame, textvariable = self.values1[i], width = 20)
            entry1.grid(row = row, column = 1, padx=2, pady =2, sticky = "w")
            self.field_stage1.append(entry1)
            row +=1
            

    def create_label2(self):
        global comp_data
        #create entry for sales quote which will be the record for the database
        label = tk.Label(self.frame, text = "CUSTOMER NAME" ,bg = "grey", font=("Helvetica", 8))
        label.grid(row = 1, column = 3 , sticky ="e")
        
        self.entry2 =tk.Entry(self.frame, width = 20)
        self.entry2.grid(row =1, column = 4, padx=2, pady =2, sticky = "w")
        self.field_stage1.append(self.entry2)
      
        label = tk.Label(self.frame, text = "SALES QT. NO." ,bg = "grey", font=("Helvetica", 8))
        label.grid(row = 2, column = 3 , sticky ="e")
        var = tk.StringVar()
        self.entry3 =tk.Entry(self.frame,textvariable = var ,width = 20)
        self.entry3.grid(row =2, column = 4, padx=2, pady =2, sticky = "w")
        self.entry3.bind("<FocusOut>", self.write_data1)
      
        
    def write_data1(self, event):
        global sq_number
        sq_number = self.entry3.get()
		#compressor_info = self.entry
        return self.check_sq_number(sq_number)


    def check_sq_number(self, sq_number):
            
            while True:
                try:
                        wb = wb = openpyxl.load_workbook(filename='example1.xlsx', data_only=True)
                        ws1 = wb.get_sheet_by_name("BASE STAGE")
                        # keep the count of row to find out which row is empty
                        self.row1 = 1
                        for row in ws1.iter_rows(row_offset = 1):
                            self.row1 +=1
                            #check if the row is empty 
                            if sq_number == row[0].value:
                                return  self.get_rowvalue_read()
                            elif row[0].value is None:
                                 return self.get_rowvalue_write(sq_number)
                except ValueError:
                        popupmsg("Enter Serial No")
    def popupmsg1(self, msg):

    #This is  popupmsg which could be used to inform the user of any event
        self.popup = tk.Tk()
        self.popup.wm_title("!")
        label = ttk.Label(self.popup, text=msg, font=NORM_FONT)
        label.pack(side="top", fill="x", pady=10)
        b1 = ttk.Button(self.popup, text="READ", command = self.calldata_from_excel)
        b2 = ttk.Button(self.popup, text="EDIT", command = self.edit_mode_on)
        b1.pack()
        b2.pack()
	
    def edit_mode_on(self):
        global isedit
        isedit = True
        return self.calldata_from_excel()

	
    def get_rowvalue_read(self):
        global row
        row = self.row1
        self.popupmsg1("Sales Quote Exist, Do you want read or Edit?")

    def calldata_from_excel(self):
        self.popup.destroy()
        base_stage_data = read_base_data1(row)
        base_stage_data1 = read_base_data2(row)
        self.write_create_label3(base_stage_data)
        self.write_create_label5(base_stage_data1)
        self.write_create_label8(base_stage_data1)
        self.write_create_label9(base_stage_data1)
        self.write_create_label10(base_stage_data1)
        if check_stages_data_stage2(row):
                self.write_stage_frame1(col=2)
                self.write_stage_frame2(col=3)
                self.write_stage_frame3(col=2)
                self.write_stage_frame4(col=2)
        print ("Found serial number at :", row)
    
    def get_rowvalue_write(self, sq_number):  
    # This function set the row number where record will be written
        
        row1 = self.row1
        print ("Write data to ", row1)
        wb = openpyxl.load_workbook(filename='example1.xlsx', data_only = True)
        ws1 = wb.get_sheet_by_name("BASE STAGE")
        ws1.cell(row = row1, column = 1).value = sq_number
        wb.save("example1.xlsx")
        return row_number(row1, sq_number)  

        
    def create_label3(self):
        left_table2 = ["COMPRESSOR MODEL","COMPRESSOR SN", "RATED HP", "RPM"]
        
        row = 6
        i =1
        self.values2 = []
        for txt3 in left_table2:
             i = len(self.values2)
             label = tk.Label(self.frame, text =txt3, bg = "grey", font = ("Helvetica",8), width = 25)
             label.grid(row = row, column = 0 , sticky ="e")
             self.values2.append(tk.StringVar())
             self.values2[i].trace("w", lambda name, index, mode, var=self.values2[i], i=i:self.get_comp_data(var, i)) 
             entry3 =tk.Entry(self.frame, textvariable= self.values2[i], width = 20)
             entry3.grid(row =row, column = 1, padx=2, pady =2, sticky = "w")
             self.field_stage1.append(entry3)
             row +=1
    
    def write_create_label3(self, base_stage_data):
    #This writes the data into the UI when the read button is called 
        entry1 =tk.Entry(self.frame, width = 20)
        entry1.insert(tk.END, base_stage_data[0])
        entry1.grid(row =1 , column = 1, padx=2, pady =2, sticky = "w")
        self.edit_data_label3.append(entry1)
        entry2 =tk.Entry(self.frame, width = 20)
        entry2.insert(tk.END, base_stage_data[1])
        entry2.grid(row =2 , column = 1, padx=2, pady =2, sticky = "w")
        self.edit_data_label3.append(entry2)
        entry2A =tk.Entry(self.frame, width = 20)
        entry2A.insert(tk.END, base_stage_data[2])
        entry2A.grid(row = 1 , column = 4, padx=2, pady =2, sticky = "w")
        self.edit_data_label3.append(entry2A)
        
        row = 6
        for field in base_stage_data[3:7]:
             entry3 =tk.Entry(self.frame, width = 20)
             entry3.insert(tk.END, field)
             entry3.grid(row =row, column = 1, padx=2, pady =2, sticky = "w")
             self.edit_data_label3.append(entry3)
             row +=1
        row = 6    
        for field in base_stage_data[7:13]:
            entry4 =tk.Entry(self.frame, width = 20)
            entry4.insert(tk.END, field)
            entry4.grid(row = row, column = 4, padx=2, pady =2, sticky = "w")
            self.edit_data_label3.append(entry4)
            row +=1
				
    def get_edit_data(self):
            data_write = [item.get() for item in self.edit_data_label3]
            print (data_write)
			
    def create_label4(self):
        right_table2 = ["DRIVER MODEL" , "DRIVER SN:", "STROKE", "CONN. ROD LENGTH"]
        row = 6
        self.values3 = []
        for txt4 in right_table2:
            i = len(self.values3)
            label =tk.Label(self.frame, text = txt4, bg = "grey", font=("Helvetica", 8), width = 20)
            label.grid(row =row, column =3, sticky ="e")
            self.values3.append(tk.StringVar())
            self.values3[i].trace("w", lambda name, index, mode, var=self.values3[i], i=i:
            self.get_comp_data(var, i)) 
            entry4 = tk.Entry(self.frame, textvariable = self.values3[i], width = 20)
            entry4.grid(row =row, column = 4, padx=2, pady =2, sticky = "w")
            self.field_stage1.append(entry4)                                                                  
            row += 1
    
    def get_comp_data(self, *args):
        global comp_data
        
        #this gets the compressor information form the user entry and trace any changes
        
        comp_data0 = [item.get() for item in self.values1]
        comp_data1 = [self.entry2.get()]
        comp_data2 = [item.get() for item in self.values2]
        comp_data3 = [item.get() for item in self.values3]
        comp_data = comp_data0 + comp_data1 + comp_data2 + comp_data3
        return compressor_data_excel(comp_data)
              
    def create_label5(self):
        left_table3 =["THROW NO.", "GAS DES.", "AMB. PRESS", "SUC PRESS",
             "DISCHARGE PRESS", "SUCTION TEMP", "DISCHARGE TEMP"]
        self.operating_data = []
        row = 4
        self.values = []
        stage_val = tk.StringVar()
        spinbox4 = tk.Spinbox(self.frame1, value = ("STAGE 1", "STAGE 2","STAGE 3","STAGE 4"),
                              textvariable = stage_val,width = 15)
        stage_val.trace("w", self.create_label8)
        spinbox4.grid(row = 3, column = 1, padx =2, pady =2, sticky = "w")
        self.operating_data.append(spinbox4)
        for txt5 in left_table3:
             i = len(self.values)
             self.values.append(tk.StringVar())
             self.values[i].trace("w", lambda name, index, mode, var=self.values[i], i=i:
                              self.press_calc_display(var, i))  
             label =tk.Label(self.frame1, text = txt5, bg = "grey",font=("Helvetica", 8),width = 15 )
             label.grid(row =row, column =0, sticky ="w")
             entry5 = tk.Entry(self.frame1,textvariable= self.values[i],width = 15)
             entry5.grid(row =row, column = 1, padx=2, pady =2, sticky = "w" )
             self.operating_data.append(entry5)
             row += 1
    
    def write_create_label5(self, base_stage_data1):
        #this function writes the compressor operatin data for base stage data
        row = 4
        self.operating_data_edit = []
        var = tk.StringVar()
        var.set(base_stage_data1[0])
        spinbox4 = tk.Spinbox(self.frame1, value = ("STAGE 1", "STAGE 2","STAGE 3","STAGE 4"),textvariable = var,width = 15)
        spinbox4.grid(row = 3, column = 1, padx =2, pady =2, sticky = "w")
        self.operating_data_edit.append(spinbox4)
        self.values = []
        for field  in base_stage_data1[1:8]:
            
            i = len(self.values)
            self.values.append(tk.StringVar())
            if isedit:
                self.values[i].trace("w", lambda name, index, mode, var=self.values[i], i=i:
                                  self.press_calc_display_edit(var, i))  
            entry5 = tk.Entry(self.frame1,textvariable = self.values[i], width = 15)
            entry5.insert(tk.END, field)
            entry5.grid(row =row, column = 1, padx=2, pady =2, sticky = "w" )
            self.operating_data_edit.append(entry5)
            row += 1
       
    def press_calc_display_edit(self, *args): 
		# calculated the pressure only if the edit button is on
        
        self.base_stage_input1 =[item.get() for item in self.operating_data_edit]
        print("edit mode", self.base_stage_input1)
		
        if len(self.base_stage_input1) > 7:
            pressure = press_calc(self.base_stage_input1)
            print ("pressure-Edit mode", pressure)
            entry81 = tk.Entry(self.frame3, width = 12)
            entry81.insert(tk.END, pressure[0])
            entry81.grid(row = 3, column = 1, padx=2, pady =2, sticky = "w")
            entry82 = tk.Entry(self.frame3, width = 12)
            entry82.insert(tk.END, pressure[1])
            entry82.grid(row = 4, column = 1, padx=2, pady =2, sticky = "w")
        else:
            pass
     
    
    def press_calc_display(self, *args):

    #calculates the pressure for inserting and also writes the data to excel by calling funtion base_stage_data
        
        self.base_stage_input =[item.get() for item in self.operating_data ]
        print ("write mode", self.base_stage_input)
        
        pressure = press_calc(self.base_stage_input)
        
        print ("pressure", pressure)
        entry81 = tk.Entry(self.frame3, width = 12)
        entry81.insert(tk.END, pressure[0])
        entry81.grid(row = 3, column = 1, padx=2, pady =2, sticky = "w")
        entry82 = tk.Entry(self.frame3, width = 12)
        entry82.insert(tk.END, pressure[1])
        entry82.grid(row = 4, column = 1, padx=2, pady =2, sticky = "w")
        self.field_stage1.append(entry82)    
        return base_stage_data1(self.base_stage_input, pressure)

    def create_label7(self):
        self.mat = tk.StringVar()
        self.mat.set("410SS-32HRC")
        options1 = ttk.Combobox(self.frame2,values = ("410SS-32HRC", "410SS-22HRC"),textvariable= self.mat)
        self.mat.trace("w",self.cal_fatigue)
        options1.grid(row =1, column =1 , padx =2, pady =2, sticky = "e"  )

        label =tk.Label(self.frame2, text = "FOS" ,bg = "grey", font=("Helvetica", 8))
        label.grid(row = 1, column =2, sticky ="e")
        entry6 =tk.Entry(self.frame2, width = 12)
        entry6.grid(row = 2, column = 1, padx=2, pady =2, sticky = "w" )

    def fos(self):
        mat = self.mat.get
        val = tk.StringVar()
        val.set(1.0)
        val.trace("w", self.cal_fatigue)
        self.entry1 = tk.Entry(self.frame2,textvariable=val, width = 10)
        self.entry1.grid(row =1, column = 3 , padx =2, pady =2, sticky = "w")
      
        label =tk.Label(self.frame2, text = "FATIGUE STRENGTH",bg = "grey", font=("Helvetica", 8))
        label.grid(row = 2, column =0, sticky ="e")
           
    def cal_fatigue(self,*args):
        mat = self.mat.get()
        self.state = []
        fos = 1
        user_entry = self.entry1.get()
        try:
                fos = float(user_entry)
                print ("FOS", fos)
        except ValueError:
                pass
        if mat == "410SS-22HRC":
            mat_strgth = mat_prop[2]
        else:
            mat_strgth = mat_prop[1]
        

        try:
            fatigue = int(round(mat_strgth/fos,0))
            entry6 =tk.Entry(self.frame2, width = 12)
            entry6.insert(0,fatigue)
            entry6.grid(row = 2, column = 1, padx=2, pady =2, sticky = "w" )
            self.field_stage1.append(entry6)
            return material_data_excel(fatigue)
        except ZeroDivisionError and ValueError:
            popupmsg("Cannot Divide by Zero")
              

    def create_label8(self, *args):
        # creates label and takes input for type of valve 
        left_table6 =[ "PD-SEAT","PD-CARRIER",
                       "SD-CARRIER"]
        base_stage =  self.operating_data[0].get()
        # add label for suction and discharge
        label10 = tk.Label(self.frame3, text =stage_suc[base_stage] , font=("Helvetica", 8), bg = "grey", width = 12)
        label10.grid(row = 1, column = 1, padx =2, pady = 2, sticky = "w")
        print (stage_dis[base_stage])
        label11 = tk.Label(self.frame3, text = stage_dis[base_stage], font=("Helvetica", 8), bg = "grey", width = 12)
        label11.grid(row = 1, column = 2, padx =2, pady = 2, sticky = "w")

        
        label =tk.Label(self.frame3, text = "VALVE-TYPE",bg = "grey", font=("Helvetica", 8))
        label.grid(row =2, column =0, sticky ="w")

        # spinbox lets you select the type of valve 
        self.val1 = tk.StringVar()
        col1 =1
        
        spinbox2 = tk.Spinbox(self.frame3, value = ("TYPE 0","TYPE 1","TYPE 2", "TYPE 3", "TYPE4"), textvariable = self.val1, width = 12,
                              command = lambda:self.pocket_stage1(self.val1.get(),col1))
        spinbox2.grid(row =2, column = 1 , padx =2, pady =2, sticky = "w"  )
        self.val2 = tk.StringVar()
       
        col2 = 2
        spinbox3 = tk.Spinbox(self.frame3, value = ("TYPE 0","TYPE 1","TYPE 2", "TYPE 3", "TYPE4"), textvariable = self.val2, width = 12,
                              command = lambda:self.pocket_stage1(self.val2.get(),col2))
        spinbox3.grid(row =2, column = 2 , padx =2, pady =2, sticky = "w"  )

        
        label =tk.Label(self.frame3, text = "PRESS DIFF",bg = "grey", font=("Helvetica", 8))
        label.grid(row = 3, column =0, sticky ="w")
        entry81 = tk.Entry(self.frame3, width = 12)
        entry81.grid(row = 3, column = 1, padx=2, pady =2, sticky = "w")
        self.field_stage1.append(entry81)

        label =tk.Label(self.frame3, text = "PRESS RATIO",bg = "grey", font=("Helvetica", 8))
        label.grid(row = 4, column =0, sticky ="w")
        entry81 = tk.Entry(self.frame3, width = 12)
        entry81.grid(row = 4, column = 1, padx=2, pady =2, sticky = "w")
        self.field_stage1.append(entry81)


        label =tk.Label(self.frame3, text = "NO.OF MODULES",bg = "grey", font=("Helvetica", 8))
        label.grid(row = 5, column =0, sticky ="w")
        
        sv1 = tk.StringVar()
        sv1.trace("w", lambda name, index, mode, sv1=sv1: write_module_data(sv1.get(), col =1))
        self.no_of_modules_suc = tk.Entry(self.frame3, textvariable = sv1, width = 12)
        self.no_of_modules_suc.grid(row = 5, column = 1, padx=2, pady =2, sticky = "w", )
        self.field_stage1.append(self.no_of_modules_suc)
        
        sv2 = tk.StringVar()
        sv2.trace("w", lambda name, index, mode, sv2=sv2: write_module_data(sv2.get(),col =2))
        self.no_of_modules_dis = tk.Entry(self.frame3,textvariable =sv2 , width = 12)
        self.no_of_modules_dis.grid(row = 5, column = 2, padx=2, pady =2, sticky = "w", )
        self.field_stage1.append(self.no_of_modules_dis)

        #enter pocket depth information here 
        row = 6
        for txt7 in left_table6:
                    
             label =tk.Label(self.frame3, text = txt7,bg = "grey", font=("Helvetica", 8))
             label.grid(row =row, column =0, sticky ="w")
             entry83 = tk.Entry(self.frame3, width = 12)
             entry83.grid(row =row, column = 1, padx=2, pady =2, sticky = "w")
             entry84 =tk.Entry(self.frame3, width = 12)
             entry84.grid(row = row, column =2, padx =2, pady =2, sticky = "w")
             self.field_stage1.append(entry83)
             row += 1
             
        
        
    def write_create_label8(self, base_stage_data1):
        # creates label and takes input for type of valve 
     
        # spinbox lets you select the type of valve
        
        label10 = tk.Label(self.frame3, text = "1S" , font=("Helvetica", 8), bg = "grey", width = 12)
        label10.grid(row = 1, column = 1, padx =2, pady = 2, sticky = "w")
    
        label11 = tk.Label(self.frame3, text = "1S", font=("Helvetica", 8), bg = "grey", width = 12)
        label11.grid(row = 1, column = 2, padx =2, pady = 2, sticky = "w")
        
        val1 = tk.StringVar()
        spinbox2 = tk.Spinbox(self.frame3, value = ("TYPE 0","TYPE 1","TYPE 2", "TYPE 3", "TYPE4"), textvariable = val1, width = 12)
        spinbox2.delete(0, "end")
        spinbox2.insert(tk.END,base_stage_data1[8] )
        spinbox2.grid(row = 2, column = 1 , padx =2, pady =2, sticky = "w"  )
        self.edit_data_label8.append(spinbox2)
        val2 = tk.StringVar()
        val2.set(base_stage_data1[12])
        spinbox3 = tk.Spinbox(self.frame3, value = ("TYPE 0","TYPE 1","TYPE 2", "TYPE 3", "TYPE4"), textvariable = val2, width = 12)
        spinbox3.delete(0, "end")
        spinbox3.insert(tk.END,base_stage_data1[15] )
        spinbox3.grid(row = 2, column = 2 , padx =2, pady =2, sticky = "w"  )
        self.edit_data_label8.append(spinbox3)
        # enter pressure difference 
        entry81 = tk.Entry(self.frame3, width = 12)
        entry81.insert(tk.END, base_stage_data1[9])
        entry81.grid(row = 3, column = 1, padx=2, pady =2, sticky = "w")
		#values are calucated here _____
        # enter pressure ratio 
        entry81 = tk.Entry(self.frame3, width = 12)
        entry81.insert(tk.END, base_stage_data1[10])
        entry81.grid(row = 4, column = 1, padx=2, pady =2, sticky = "w")
		#values are calucated here ___
        #enter no of modules suction
        
        no_of_modules_suc = tk.Entry(self.frame3, width = 12)
        no_of_modules_suc.insert(tk.END, base_stage_data1[11])
        no_of_modules_suc.grid(row = 5, column = 1, padx=2, pady =2, sticky = "w", )
        self.edit_data_label8.append(no_of_modules_suc)
        #enter no of modules discharge 
        no_of_modules_dis = tk.Entry(self.frame3, width = 12)
        no_of_modules_dis.insert(tk.END, base_stage_data1[16])
        no_of_modules_dis.grid(row = 5, column = 2, padx=2, pady =2, sticky = "w", )
        self.edit_data_label8.append(no_of_modules_dis)
        row = 6
    #enter pocket depth for suction 
        for txt7 in base_stage_data1[12:15]:
          
             entry83 = tk.Entry(self.frame3, width = 12)
             entry83.insert(tk.END, txt7)
             entry83.grid(row =row, column = 1, padx=2, pady =2, sticky = "w")
             self.edit_data_label8.append(entry83)
             row += 1 
             
        row = 6
        #enter pocket depth for discharge
        for txt7 in base_stage_data1[17:20]:
             entry84 =tk.Entry(self.frame3, width = 12)
             entry84.insert(tk.END, txt7)
             entry84.grid(row = row, column =2, padx =2, pady =2, sticky = "w")
             self.edit_data_label8.append(entry84)
             row += 1 
             
    def pocket_stage1(self, valve_type, col):
        global modules_suc, modules_dis
        self.seat_pocket_value1 =[]
        self.seat_pocket_value2 =[]
        self.seat_dim =[0,0,0]
        print (valve_type)
        print ("COLUMN AT POCKET DIM ", col)
        modules_suc = self.no_of_modules_suc.get()
        
        modules_dis = self.no_of_modules_dis.get()
       
        if valve_type == "TYPE 1":
            self.seat_dim[0] = 0
            self.seat_dim[1] = 14.92
            self.seat_dim[2] = 18
        elif valve_type == "TYPE 2":
            self.seat_dim[0] = 14.92
            self.seat_dim[1]= 0
            self.seat_dim[2] = 18
        elif valve_type == "TYPE 3":
            self.seat_dim[0] = 7.92
            self.seat_dim[1]= 7.0
            self.seat_dim[2] = 18
       
        print ("Seat Pocket Depth ", self.seat_dim)
        row = 6
        self.pocket_value1 = []
        
        for item in self.seat_dim:
            i = len(self.pocket_value1)
            self.pocket_value1.append(tk.StringVar())
            self.pocket_value1[i].trace("w",lambda name, index, mode, var=self.pocket_value1[i], i=i:
                              self.entryUpdate(var, i, col))
       
            entry84 = tk.Entry(self.frame3, textvariable = self.pocket_value1[i], width = 10)
            entry84.insert(tk.END, item)
            entry84.grid(row = row, column = col, padx=2, pady =2, sticky = "w")
            row +=1
            
        return write_valve_type(valve_type, col,)
        
            
        
    def entryUpdate(self, var, i , col):
        global pocket_depth
        
        print ("COLUMN NUMBER", col)
        #gets the data from the pocket_data depending on type of valve
        if col == 1:
                pocket_depth = [item.get()for item in self.pocket_value1]
                # get the pocket depth and write data to excel for suction
                print ("POCKET DIM SUC", pocket_depth)
                return write_pocket_depth(pocket_depth, col)
        elif col == 2:
                pocket_depth = [item.get()for item in self.pocket_value1]
                # get the pocket depth and write data to excel for suction
                print ("POCKET DIM DIS", pocket_depth)
                return write_pocket_depth(pocket_depth, col)

    def pocket_depth_value1(self):
        global pocket_depth
        #  calculates the pocket from the input list 
        if len(pocket_depth) == 3:
            seat_pocket_depth     = float(pocket_depth[0])
            carrier_pocket_depth  = float(pocket_depth[1])
            carrier_slot_depth    = float(pocket_depth[2])
            valve_pocket = [seat_pocket_depth, carrier_pocket_depth, carrier_slot_depth]
            return valve_pocket 
        else:
                pass
   
    def create_label9(self):
        seat_prop = [ "MIN SEAT THICKNESS(IN)", "DESIGN SEAT THICKNESS(IN)",
                      "DESIGN PRESS.(PSI)", "DEFLECTION",
                      "NATURAL FREQ", "FOS", "VALVE MAT.", "MODULE MAT."]
        self.seat_design_values = []
        row = 2
        self.values4 = []
        self.values5 = []
        col = 2
        for prop in seat_prop:
            i = len(self.values4)
            j = len(self.values5)
            self.values4.append(tk.StringVar())
            self.values5.append(tk.StringVar())
            self.values4[i].trace("w", lambda name, index, mode, var=self.values4[i], i=i:
                              self.get_seat_thickness1(col= 1))
            self.values5[j].trace("w", lambda name, index, mode, var=self.values5[j], j=j:
                              self.get_seat_thickness1(col = 2))
            
            label = tk.Label(self.frame4, text = prop, font=("Helvetica", 8), bg = "grey", width = 23)
            label.grid(row = row, column = 1, padx =2, pady = 2, sticky = "w")

            entry9 = tk.Entry(self.frame4,textvariable = self.values4[i],  width = 12)
            entry9.grid(row =row, column = col, padx=2, pady =2, sticky = "w" )

            entry10 = tk.Entry(self.frame4,textvariable = self.values5[i], width = 12)
            entry10.grid(row =row, column = col+1, padx =2, pady =2, sticky = "w")
            if col == 2:    
                    self.seat_design_values.append(entry9)
            elif col == 3:
                    self.seat_design_values.append(entry10) 
            row +=1
            
            
    def write_create_label9(self, base_stage_data1):
       
        row = 2
        self.values40 = []
		
        for seat_values_suc in base_stage_data1[20:28]:
            i = len(self.values40)
            self.values40.append(tk.StringVar())
            if isedit:
                self.values40[i].trace("w", lambda name, index, mode, var=self.values40[i], i=i:self.get_seat_thickness1_edit(col= 1))
                       
            entry9 = tk.Entry(self.frame4, textvariable = self.values40[i], width = 12)
            entry9.insert(tk.END, seat_values_suc)
            entry9.grid(row =row, column = 2, padx=2, pady =2, sticky = "w" )
            self.edit_data_label9.append(entry9)
            row +=1
        
        row = 2
        self.values50 =[]
        for seat_values_dis in base_stage_data1[28:36]:
            j = len(self.values50)
            self.values5.append(tk.StringVar())
            if isedit:
                    self.values50[j].trace("w", lambda name, index, mode, var=self.values50[j],

                                                       j=j:self.get_seat_thickness1_edit(col = 2))
					
            entry10 = tk.Entry(self.frame4,textvariable = self.values50[i], width = 12)
            entry10.insert(tk.END, seat_values_dis)
            entry10.grid(row =row, column = 3, padx =2, pady =2, sticky = "w")
            self.edit_data_label9.append(entry10)
            row +=1  
    
	
    def get_seat_thickness1_edit(self, col):
        print("column number", col)
        if col == 1:
             seat_suc = [item.get() for item in self.values4]
             print ("SEAT SUC", seat_suc, "COL #", col)
             return self.display_valve_data(seat_suc, col)
        elif col ==2:    
             seat_dis = [item.get() for item in self.values5]
             print ("SEAT DIS", seat_dis, "COL #", col )
             return self.display_valve_data(seat_dis, col)
	
    
    def get_seat_thickness1(self, col):
        print("column number", col)
        if col == 1:
             seat_suc = [item.get() for item in self.values4]
             print ("SEAT SUC", seat_suc, "COL #", col)
             return self.display_valve_data(seat_suc, col),base_stage_data2(seat_suc, col)
        elif col ==2:    
             seat_dis = [item.get() for item in self.values5]
             print ("SEAT DIS", seat_dis, "COL #", col )
             return self.display_valve_data(seat_dis, col), base_stage_data2(seat_dis,col)
        
            
    def display_valve_data(self, seat, col ):
        global pocket_depth, seat_thick, carrier_thick, valve_thick
        print ("column at frame 5", col)
        if col == 1:
        #suction valve dimensions
                pocket_depth = self.pocket_depth_value1()
                print ("Pocket_depth entry", pocket_depth)
                try:
                    design_seat_thick_suc = float(seat[1])
                    print ("Design Seat Thickness -SUC", design_seat_thick_suc)
                    seat_pocket_depth_suc = float(pocket_depth[0])/25.4     
                    print ("SEAT POCKET SUC ", seat_pocket_depth_suc)
                    seat_thick = seat_pocket_depth_suc + design_seat_thick_suc
                    seat_thick = round(seat_thick, 3)
                    print ( "TOT SEAT THICK SUC ", seat_thick)
                    carrier_pocket = pocket_depth[1]/25.4
                    carrier_slot = pocket_depth[2]/25.4
                    carrier_thick = carrier_pocket + carrier_slot
                    carrier_thick = round(carrier_thick, 3)
                    print("Carrier thickness", carrier_thick)
                    valve_thick = carrier_thick + seat_thick
                    valve_thick = round(valve_thick, 3)
                    entry12 = tk.Entry(self.frame5, width = 15)
                    entry13 = tk.Entry(self.frame5, width = 15)
                    entry14 = tk.Entry(self.frame5, width = 15)
                    entry12.insert(tk.END, seat_thick)
                    entry13.insert(tk.END, carrier_thick)
                    entry14.insert(tk.END, valve_thick)
                    entry12.grid(row = 2,column = 1, padx=4, pady =2, sticky = "we")
                    entry13.grid(row = 3,column = 1, padx=4, pady =2, sticky = "we")
                    entry14.grid(row = 4,column = 1, padx=4, pady =2, sticky = "we")
                    self.field_stage1.append(entry12)
                    self.field_stage1.append(entry13)
                    self.field_stage1.append(entry14)                                                        
                    return valve_data(col)
                except ValueError:
                            pass
        #discharge valve dimensions 
        elif col  == 2:
                pocket_depth = self.pocket_depth_value1()
                try:
                    design_seat_thick_dis = float(seat[1])
                    print ("Design Seat Thickness - DIS", design_seat_thick_dis)
                    seat_pocket_depth = float(pocket_depth[0])/25.4  
                    print ("SEAT POCKET DIS ", seat_pocket_depth)
                    seat_thick = seat_pocket_depth + design_seat_thick_dis
                    seat_thick = round(seat_thick, 3)
                    print ( "TOT SEAT THICK DIS", seat_thick)
                    carrier_pocket = pocket_depth[1]/25.4
                    carrier_slot = pocket_depth[2]/25.4
                    carrier_thick = carrier_pocket + carrier_slot
                    carrier_thick = round(carrier_thick, 3)
                    valve_thick = carrier_thick + seat_thick
                    valve_thick = round(valve_thick, 3)
                    entry12 = tk.Entry(self.frame5, width = 12)
                    entry13 = tk.Entry(self.frame5, width = 12)
                    entry14 = tk.Entry(self.frame5, width = 12)
                    entry12.insert(tk.END, seat_thick)
                    entry13.insert(tk.END, carrier_thick)
                    entry14.insert(tk.END, valve_thick)
                    entry12.grid(row = 2,column = 2, padx=4, pady =2, sticky = "we")
                    entry13.grid(row = 3,column = 2, padx=4, pady =2, sticky = "we")
                    entry14.grid(row = 4,column = 2, padx=4, pady =2, sticky = "we")
                    self.field_stage1.append(entry12)
                    self.field_stage1.append(entry13)
                    self.field_stage1.append(entry14) 
                    return valve_data(col)
                except ValueError:
                                pass    
        

       
    def create_label10(self, *args):


        valve_dim = [ "SEAT THICKNESS (IN)", "CARRIER THICKNESS(IN)",
                      "VALVE THICKNESS (IN)" ]
        row = 2
        

        for prop in valve_dim:
            label = tk.Label(self.frame5, text = prop, font=("Helvetica", 8), bg = "grey", width = 23)
            label.grid(row = row, column = 0, padx =2, pady = 2, sticky = "w")
            entry9 = tk.Entry(self.frame5, width = 12)
            entry9.grid(row =row, column = 1, padx=2, pady =2, sticky = "w" )
            entry10 = tk.Entry(self.frame5, width = 12)
            entry10.grid(row =row, column = 2, padx =2, pady =2, sticky = "w")
            self.field_stage1.append(entry9)
            row +=1

    def write_create_label10(self, base_stage_data1):


        row = 2
        

        for dim1 in base_stage_data1[36:39]:
            
            entry9 = tk.Entry(self.frame5, width = 12)
            entry9.insert(tk.END, dim1 )
            entry9.grid(row =row, column = 1, padx=2, pady =2, sticky = "w" )
            row +=1
            
        row = 2
        for dim2 in base_stage_data1[39:43]:    
            entry10 = tk.Entry(self.frame5, width = 12)
            entry10.insert(tk.END, dim2)
            entry10.grid(row =row, column = 2, padx =2, pady =2, sticky = "w")
            row +=1
 
    
    def display_stage_frame1(self, col):
        #This displays 2 stages entry fileds when stage 2 is selected
        global input_col_2, input_col_3, input_col_4
        print ("Column Number", col)
        if col == 2:
                col_1 = 3
        elif col == 3:
                col_1 = 5
        elif col ==4:
                col_1 = 7
        elif col == 5:
                col_1 = 9
        stage_val = tk.StringVar()
        
        # This label displays the stages and stores in the class field to be fetched for deleting and writing data to excel sheet
        spinbox4 = tk.Spinbox(self.frame1, value = ("STAGE 1", "STAGE 2","STAGE 3","STAGE 4"),
                              textvariable = stage_val,width = 15, command = lambda: self.display_stage_frame2(stage_val.get(),col))
        spinbox4.grid(row = 3, column = col, padx =2, pady =2, sticky = "w")
        
         
        if col == 2:
                self.field_stage2_oper.append(spinbox4)   
        elif col == 3:
                self.field_stage3_oper.append(spinbox4)
               
        elif col == 4:
                self.field_stage4_oper.append(spinbox4)
              
        elif col == 5:
                self.field_stage5_oper.append(spinbox4)

        elif col == 6:
                self.field_stage6_oper.append(spinbox4)
                        
               
        self.valve_oper = []
        for row in self.rows1:
            i = len(self.valve_oper)
            self.valve_oper.append(tk.StringVar())
            self.valve_oper[i].trace("w", lambda name, index, mode, var=self.valve_oper[i], i=i:
                              self.get_data1_frame1(var, i, col,))
            
            entry10 =tk.Entry(self.frame1, textvariable = self.valve_oper[i], width = 15)
            entry10.grid(row = row , column = col, padx=4, pady =2, sticky = "we" )
            if col == 2:
                self.field_stage2_oper.append(entry10)
            elif col == 3:
                self.field_stage3_oper.append(entry10)
            elif col == 4:
                self.field_stage4_oper.append(entry10)
            elif col == 5:
                self.field_stage5_oper.append(entry10)
            elif col == 6:
                self.field_stage5_oper.append(entry10) 
	
    def write_stage_frame1(self, col):
        #This displays 2 stages entry fileds when stage 2 is selected
        stage = str(stage_data_col2[0])
        stage_val = tk.StringVar()
        #tage_val.set(stage)
        # This label displays the stages and stores in the class field to be fetched for deleting and writing data to excel sheet
        spinbox4 = tk.Spinbox(self.frame1, value = ("STAGE 1", "STAGE 2","STAGE 3","STAGE 4"), textvariable = stage_val, width = 15)
        spinbox4.delete(0, "end")
        spinbox4.insert(tk.END, stage)
        spinbox4.grid(row = 3, column = col, padx =2, pady =2, sticky = "w")
        
        row = 4
        for item  in stage_data_col2[1:8]:
            entry10 =tk.Entry(self.frame1,width = 15)
            entry10.insert(tk.END, item)
            entry10.grid(row = row , column = col, padx=4, pady =2, sticky = "we" )
            row += 1
        
    def get_data1_frame1(self,sv, i, col):
        print (sv, i, sv.get())
        global input_col_2, input_col_3, input_col_4, input_col_5, input_col_6
        if col == 2:
            input_col_2 = [item.get() for item in self.field_stage2_oper]
        elif col ==3:
            input_col_3 = [item.get() for item in self.field_stage3_oper]
        elif col == 4:
            input_col_4 =  [item.get() for item in self.field_stage3_oper]
        elif col == 5:
            input_col_5 =  [item.get() for item in self.field_stage3_oper]
        elif col == 6:
            input_col_6 = [item.get() for item in self.field_stage3_oper]
        elif col == 7:
            input_col_7 = [item.get() for item in self.field_stage3_oper] 
        elif col == 8:
            input_col_8 = [item.get() for item in self.field_stage3_oper]       
       

        print (input_col_2)
        print (input_col_3)
        print (input_col_4)
        print (input_col_5)
        return self.display_stage_frame2_cal(col)     

    def display_stage_frame2_cal(self, col):
        """ This function takes the values user enters and calculates the pressure differnece
           and pressure ratio and displays it """
        print ("COlUMN #", col)
        global input_col_2, input_col_3, input_col_4, input_col_5, input_col_6
       
        print ("INPUT col 2 lst", input_col_2)
        print (input_col_3)
        print (input_col_4)
        print (input_col_5)

       
        if col == 2:
                col_1  = 3
                pressure = press_calc(input_col_2)
                stage_info = input_col_2
        elif col == 3:
                col_1 = 5 
                pressure = press_calc(input_col_3)
                stage_info = input_col_3
        elif col  == 4:
                col_1 = 7
                pressure = press_calc(input_col_4)
                stage_info = input_col_4
        elif col == 5:
                col_1 = 9
                pressure = press_calc(input_col_4)
                stage_info= input_col_5
        elif col == 6:
                col_1 = 11
                pressure = press_calc(input_col_4)
                stage_info= input_col_6        
                        
               
        entry11 = tk.Entry(self.frame3, width = 12)
        entry11.insert(tk.END, pressure[0])
        entry11.grid(row = 3, column = col_1, padx=4, pady =2, sticky = "we")

        entry13 = tk.Entry(self.frame3, width = 12)
        entry13.insert(tk.END, pressure[1])
        entry13.grid(row = 4, column = col_1 , padx =4, pady =2, sticky = "we")
        return base_stage_data1(stage_info, pressure)

    
    
    def display_stage_frame2(self, stage, col):
         global input_col_2, input_col_3, input_col_4, input_col_5 
         print ("Stage Number", stage)
         print ("INPUT FROM COLUMN", input_col_2)
         print ("column number: " , col)
            
      
         if col == 2:
                col_1 = 3
         elif col == 3:
                col_1 = 5
         elif col == 4:
                col_1 = 7
         elif col == 5:
                col_1 = 9
        
         #dictionary for stage and suction and discharge
         stage_suc ={ "STAGE 1" :"1S", "STAGE 2":"2S", "STAGE 3": "3S", "STAGE 4": "4S"}
         stage_dis ={ "STAGE 1" :"1D", "STAGE 2":"2D", "STAGE 3": "3D", "STAGE 4": "4D"}

         #adds label to frame 3 where suction and discharge values for the valve is stored
         print (stage_suc[stage])
         label10 = tk.Label(self.frame3, text =stage_suc[stage] , font=("Helvetica", 8), bg = "grey", width = 12)
         label10.grid(row = 1, column = col_1, padx =2, pady = 2, sticky = "w")
         print (stage_dis[stage])
         label11 = tk.Label(self.frame3, text = stage_dis[stage], font=("Helvetica", 8), bg = "grey", width = 12)
         label11.grid(row = 1, column = col_1+1, padx =2, pady = 2, sticky = "w")

         #depending of the col number the label will be stored for deleting 
         if col == 2:
                self.label_1.append(label10)
                self.label_1.append(label11)
         elif col == 3:
                self.label_2.append(label10)
                self.label_2.append(label11) 
         elif col == 4:
                self.label_3.append(label10)
                self.label_3.append(label11) 
         elif col == 5:
                self.label_4.append(label10)
                self.label_4.append(label11)

         self.label_1.append(label11)
         val1 = tk.StringVar()
        # This box lets you select the type of valve and then ouputs the pocket depth and slot depth 
         spinbox3 = tk.Spinbox(self.frame3, value = ("TYPE 0", "TYPE 1","TYPE 2", "TYPE 3", "TYPE 4"), textvariable = val1,width =12,
                              command = lambda:self.pocket_stage(val1.get(), col_1))
         spinbox3.grid(row = 2, column = col_1 , padx =2, pady =2, sticky = "we",)   

         val2 = tk.StringVar()       
         spinbox4 = tk.Spinbox(self.frame3, value = ("TYPE 0", "TYPE 1","TYPE 2", "TYPE 3", "TYPE 4"), textvariable = val2,width =12,
                              command = lambda:self.pocket_stage(val2.get(), col_1+1))
         spinbox4.grid(row = 2, column = col_1+1 , padx =2, pady =2, sticky = "we",)  
         
         

         if col == 2:
                self.field_stage2_pd.append(spinbox3)
                self.field_stage2_pd.append(spinbox4)
         elif col == 3:
                self.field_stage3_pd.append(spinbox3)
                self.field_stage3_pd.append(spinbox4)
         elif col == 4:
                self.field_stage4_pd.append(spinbox3)
                self.field_stage4_pd.append(spinbox4)
         elif col == 5:
                self.field_stage5_pd.append(spinbox3)
                self.field_stage5_pd.append(spinbox4)



                
        #check the first entry of inp_col_2:
        #depending on the stage number  put labels 1S, 1D
    

         for row in self.rows2:

            entry11 = tk.Entry(self.frame3, width = 12)
            entry11.grid(row = row, column = col_1, padx=4, pady =2, sticky = "we")

            entry13 = tk.Entry(self.frame3, width = 12)
            entry13.grid(row =row, column = col_1 +1 , padx =4, pady =2, sticky = "we")
                           
            if col == 2:
                self.field_stage2_pd.append(entry11)
                self.field_stage2_pd.append(entry13)
            elif col == 3:
                self.field_stage3_pd.append(entry11)
                self.field_stage3_pd.append(entry13)
            elif col == 4:
                self.field_stage4_pd.append(entry11)
                self.field_stage4_pd.append(entry13)
            elif col == 5:
                self.field_stage5_pd.append(entry10)
                self.field_stage5_pd.append(entry13)

         return self.display_stage_frame3(col) 


		 
    def write_stage_frame2(self, col):
         stage = stage_data_col2[0]
         col_1 = col
         #dictionary for stage and suction and discharge
         stage_suc ={ "STAGE 1" :"1S", "STAGE 2":"2S", "STAGE 3": "3S", "STAGE 4": "4S"}
         stage_dis ={ "STAGE 1" :"1D", "STAGE 2":"2D", "STAGE 3": "3D", "STAGE 4": "4D"}
         #adds label to frame 3 where suction and discharge values for the valve is stored
         print (stage_suc[stage])
         label10 = tk.Label(self.frame3, text =stage_suc[stage] , font=("Helvetica", 8), bg = "grey", width = 12)
         label10.grid(row = 1, column = col_1, padx =2, pady = 2, sticky = "w")
         print (stage_dis[stage])
         label11 = tk.Label(self.frame3, text = stage_dis[stage], font=("Helvetica", 8), bg = "grey", width = 12)
         label11.grid(row = 1, column = col_1+1, padx =2, pady = 2, sticky = "w")

         #depending of the col number the label will be stored for deleting 
         if col == 2:
                self.label_1.append(label10)
                self.label_1.append(label11)
         elif col == 3:
                self.label_2.append(label10)
                self.label_2.append(label11) 
         elif col == 4:
                self.label_3.append(label10)
                self.label_3.append(label11) 
         elif col == 5:
                self.label_4.append(label10)
                self.label_4.append(label11)

         self.label_1.append(label11)
         val1 = tk.StringVar()
        # This box lets you select the type of valve and then ouputs the pocket depth and slot depth 
         spinbox3 = tk.Spinbox(self.frame3, value = ("TYPE 0", "TYPE 1","TYPE 2", "TYPE 3", "TYPE 4"), textvariable = val1,width =12)
         spinbox3.delete(0, "end")
         spinbox3.insert(0, stage_data_col2[8])
         spinbox3.grid(row = 2, column = col_1 , padx =2, pady =2, sticky = "we",)   

         val2 = tk.StringVar()       
         spinbox4 = tk.Spinbox(self.frame3, value = ("TYPE 0", "TYPE 1","TYPE 2", "TYPE 3", "TYPE 4"), textvariable = val2,width =12)
         spinbox4.delete(0, "end")
         spinbox4.insert(0, stage_data_col2[15])					  
         spinbox4.grid(row = 2, column = col_1+1 , padx =2, pady =2, sticky = "we",)  
         
         

         if col == 2:
                self.field_stage2_pd.append(spinbox3)
                self.field_stage2_pd.append(spinbox4)
         elif col == 3:
                self.field_stage3_pd.append(spinbox3)
                self.field_stage3_pd.append(spinbox4)
         elif col == 4:
                self.field_stage4_pd.append(spinbox3)
                self.field_stage4_pd.append(spinbox4)
         elif col == 5:
                self.field_stage5_pd.append(spinbox3)
                self.field_stage5_pd.append(spinbox4)



                
        #check the first entry of inp_col_2:
        #depending on the stage number  put labels 1S, 1D
    
         row = 3
         for item  in stage_data_col2[9:15]:

            entry11 = tk.Entry(self.frame3, width = 12)
            entry11.insert(tk.END, item)
            entry11.grid(row = row, column = col_1, padx=4, pady =2, sticky = "we")
            row += 1
         
         row = 5 
         for item  in stage_data_col2[16:20]:	
            entry13 = tk.Entry(self.frame3, width = 12)
            entry13.insert(tk.END, item)
            entry13.grid(row =row, column = col_1 +1 , padx =4, pady =2, sticky = "we")
            row += 1
      	
            if col == 2:
                self.field_stage2_pd.append(entry11)
                self.field_stage2_pd.append(entry13)
            elif col == 3:
                self.field_stage3_pd.append(entry11)
                self.field_stage3_pd.append(entry13)
            elif col == 4:
                self.field_stage4_pd.append(entry11)
                self.field_stage4_pd.append(entry13)
            elif col == 5:
                self.field_stage5_pd.append(entry10)
                self.field_stage5_pd.append(entry13)

				
    def display_stage_frame3(self,col):
        
         if col == 2:
                col_1 = 4
         elif col == 3:
                col_1 = 6
         elif col ==4:
                col_1 = 8
         elif col == 5:
                col_1 = 10    
         self.values6 = []
         self.values7 = []   
         for row in self.rows3:
            i = len(self.values6)
            j = len(self.values7)
            self.values6.append(tk.StringVar())
            self.values7.append(tk.StringVar())
            self.values6[i].trace("w", lambda name, index, mode, var=self.values6[i], i=i:self.get_seat_entry(col_1))
            self.values7[j].trace("w", lambda name, index, mode, var=self.values7[j], j=j:self.get_seat_entry(col_1+1))
         
            entry12 = tk.Entry(self.frame4, textvariable = self.values6[i], width = 15)
            entry12.grid(row = row, column = col_1, padx=4, pady =2, sticky = "we")
          
            entry14 = tk.Entry(self.frame4, textvariable = self.values7[j], width = 15)
            entry14.grid(row = row, column = col_1+1, padx=4, pady =2, sticky = "we")
              
           
            if col == 2:
                self.field_stage2_sd.append(entry12)
                self.field_stage2_sd.append(entry14)
            elif col == 3:
                self.field_stage3_sd.append(entry12)
                self.field_stage3_sd.append(entry14)
            elif col == 4:
                self.field_stage4_sd.append(entry12)
                self.field_stage4_sd.append(entry14)
            elif col == 5:
                self.field_stage5_sd.append(entry12)
                self.field_stage5_sd.append(entry14)    

    def write_stage_frame3(self,col):
        
        if col == 2:
            col_1 = 4
        elif col == 3:
            col_1 = 6
        elif col ==4:
            col_1 = 8
        elif col == 5:
            col_1 = 10    
        row = 2
        for item in stage_data_col2[20:28]:
                entry12 = tk.Entry(self.frame4, width = 15) 
                entry12.insert(tk.END, item)
                entry12.grid(row = row, column = col_1, padx=4, pady =2, sticky = "we")
                row +=1
                
        row = 2		
        for item in stage_data_col2[28:36]:
            entry14 = tk.Entry(self.frame4, width = 15)
            entry14.insert(tk.END, item)
            entry14.grid(row = row, column = col_1+1, padx=4, pady =2, sticky = "we")
            row +=1
           
        if col == 2:
            self.field_stage2_sd.append(entry12)
            self.field_stage2_sd.append(entry14)
        elif col == 3:
            self.field_stage3_sd.append(entry12)
            self.field_stage3_sd.append(entry14)
        elif col == 4:
            self.field_stage4_sd.append(entry12)
            self.field_stage4_sd.append(entry14)
        elif col == 5:
            self.field_stage5_sd.append(entry12)
            self.field_stage5_sd.append(entry14)    




    def get_seat_entry(self, col1):
       # tracks the entry for seat design parameter and calls the fucntion to write data to excel 
         if col1 % 2 == 0:
                 seat_suc = [item.get() for item in self.values6]
                 print ("SEAT SUC", seat_suc, "COL #", col1)
                 return self.seat_thickness(seat_suc, col1), stage_data2(seat_suc, col1)
         else:
                 seat_dis = [item.get() for item in self.values7]
                 print ("SEAT DIS", seat_dis, "COL #", col1 )
                 return self.seat_thickness(seat_dis, col1), stage_data2(seat_dis, col1)
         
                 
                 
    
    def seat_thickness(self, seat_design, col ):
        seat_thickness = seat_design[1]
        print (seat_thick)
        return  self.display_stage_frame4(seat_thickness, col)

    def display_stage_frame4(self, seat_thickness, col):
         seat_thick = 0
         carrier_thick = 0
         valve_thick  = 0

         if col % 2 == 0:
                    try:
                        design_seat_thick_suc = float(seat_thickness)
                        print ("Design Seat Thickness -SUC", design_seat_thick_suc)
                        pocket_depth = self.pocket_depth_value()
                        seat_pocket_depth_suc = pocket_depth[0]/25.4
                        print ("SEAT POCKET SUC ", seat_pocket_depth_suc)
                        seat_thick = seat_pocket_depth_suc + design_seat_thick_suc
                        seat_thick = round(seat_thick, 3)
                        print ( "TOT SEAT THICK SUC ", seat_thick)
                        carrier_pocket = pocket_depth[1]/25.4
                        carrier_slot = pocket_depth[2]/25.4
                        carrier_thick = carrier_pocket + carrier_slot
                        carrier_thick = round(carrier_thick , 3)
                        valve_thick = seat_thick + carrier_thick
                        valve_thick = round(valve_thick , 3)
                        entry12 = tk.Entry(self.frame5, width = 15)
                        entry13 = tk.Entry(self.frame5, width = 15)
                        entry14 = tk.Entry(self.frame5, width = 15)
                        entry12.insert(tk.END, seat_thick)
                        entry13.insert(tk.END, carrier_thick)
                        entry14.insert(tk.END, valve_thick)
                        entry12.grid(row = 2,column = col, padx=4, pady =2, sticky = "we")
                        entry13.grid(row = 3,column = col, padx=4, pady =2, sticky = "we")
                        entry14.grid(row = 4,column = col, padx=4, pady =2, sticky = "we")
                        if col == 4:
                                self.field_stage2_vd.append(entry12)
                                self.field_stage2_vd.append(entry13)
                                self.field_stage2_vd.append(entry14)                                                        
                        elif col == 6:
                                self.field_stage3_vd.append(entry12)
                                self.field_stage3_vd.append(entry13)
                                self.field_stage3.append(entry14)
                        elif col == 8:
                                self.field_stage4_vd.append(entry12)
                                self.field_stage4_vd.append(entry13)
                                self.field_stage4_vd.append(entry14) 
                        return valve_col_data(seat_thick, carrier_thick, valve_thick, col)        
                
                    except ValueError:
                                pass

         else:
             
                    try:                
                        design_seat_thick_dis = float(seat_thickness)
                        print ("Design Seat Thickness - DIS", design_seat_thick_dis)
                        pocket_depth = self.pocket_depth_value()
                        seat_pocket_depth_dis = pocket_depth[0]/25.4
                        carrier_pocket = pocket_depth[1]/25.4
                        carrier_slot = pocket_depth[2]/25.4
                        print ("SEAT POCKET DIS ", seat_pocket_depth_dis)
                        seat_thick = seat_pocket_depth_dis + design_seat_thick_dis
                        seat_thick = round(seat_thick, 3)
                        carrier_thick = carrier_pocket +  carrier_slot
                        carrier_thick = round(carrier_thick, 3)
                        valve_thick =  seat_thick + carrier_thick
                        valve_thick = round(valve_thick, 3)
                        print ( "TOT SEAT THICK DIS", seat_thick)
                        entry12 = tk.Entry(self.frame5, width = 15)
                        entry13 = tk.Entry(self.frame5, width = 15)
                        entry14 = tk.Entry(self.frame5, width = 15)
                        entry12.insert(tk.END, seat_thick)
                        entry13.insert(tk.END, carrier_thick)
                        entry14.insert(tk.END, valve_thick)
                        entry12.grid(row = 2,column = col, padx=4, pady =2, sticky = "we")
                        entry13.grid(row = 3,column = col, padx=4, pady =2, sticky = "we")
                        entry14.grid(row = 4,column = col, padx=4, pady =2, sticky = "we")
                        if col == 5:
                                self.field_stage2.append(entry12)
                                self.field_stage2.append(entry13)
                                self.field_stage2.append(entry14)                                                        
                        elif col == 7:
                                self.field_stage3.append(entry12)
                                self.field_stage3.append(entry13)
                                self.field_stage3.append(entry14)
                        elif col == 9:
                                self.field_stage4.append(entry12)
                                self.field_stage4.append(entry13)
                                self.field_stage4.append(entry14)
                        return valve_col_data(seat_thick, carrier_thick, valve_thick, col)  

                    except ValueError:
                                pass    
         
    def write_stage_frame4(self, col):
        if col == 2:
            col1 = 4
        elif col == 3:
            col1 = 6
        elif col ==4:
            col1 = 8
        elif col == 5:
            col1 = 10   

        row = 2
        

        for dim1 in stage_data_col2[36:39]:
            
            entry9 = tk.Entry(self.frame5, width = 12)
            entry9.insert(tk.END, dim1 )
            entry9.grid(row =row, column = col1, padx=2, pady =2, sticky = "w" )
            row +=1
            
        row = 2
        for dim2 in stage_data_col2[39:43]:    
            entry10 = tk.Entry(self.frame5, width = 12)
            entry10.insert(tk.END, dim2)
            entry10.grid(row =row, column = col1+1, padx =2, pady =2, sticky = "w")
            row +=1     
                  
    def pocket_stage(self, valve_type, col):
        seat_dim =[0,0,0]
        
        if valve_type == "TYPE 1":
            seat_dim[0] = 0
            seat_dim[1] = 14.92
            seat_dim[2] = 18
            
        elif valve_type == "TYPE 2":
            seat_dim[0] = 14.92
            seat_dim[1]= 0
            seat_dim[2] = 18
        elif valve_type == "TYPE 3":
            seat_dim[0] = 7.92
            seat_dim[1]= 7.0
            seat_dim[2] = 18
       
        col_1 = col
        self.pocket_value2 = []  
        row = 6
        for item in seat_dim:
            i = len(self.pocket_value2)
            self.pocket_value2.append(tk.StringVar())
            self.pocket_value2[i].trace("w",lambda name, index, mode, var=self.pocket_value2[i], i=i:
                              self.get_pocket_depth(col))
       
            entry84 = tk.Entry(self.frame3, textvariable = self.pocket_value2[i], width = 10)
            entry84.insert(tk.END, item)
            entry84.grid(row = row, column = col, padx=2, pady =2, sticky = "w")
           
            if col == 3:
                self.field_stage2.append(entry84)
            elif col == 4:    
                self.field_stage2.append(entry84)
            elif col == 5:     
                self.field_stage3.append(entry84)
            elif col == 6:             
                self.field_stage3.append(entry84)
            elif col == 7:
                self.field_stage4.append(entry84)      
            elif col == 8:
                self.field_stage4.append(entry84) 
            row +=1      
            
    def get_pocket_depth(self, col):
        global pocket_depth2
        #gets the data from the pocket_data depending on type of valve
        if col%2 != 0:
                pocket_depth2 = [item.get()for item in self.pocket_value2]
                print ("POCKET DIM SUC", pocket_depth2)
                return write_pocket_depth(pocket_depth2, col)
                
        else:
                pocket_depth2 = [item.get()for item in self.pocket_value2]
                print ("POCKET DIM DIS", pocket_depth2)
                return write_pocket_depth(pocket_depth2, col)

    def pocket_depth_value(self):
        # when all three pocket depth is entered , the data is used to calculate pocket depth for valve
        if len(pocket_depth2) == 3:
                seat_pocket_depth     = float(pocket_depth2[0])
                carrier_pocket_depth  = float(pocket_depth2[1])
                carrier_slot_depth    = float(pocket_depth2[2])
                valve_pocket = [seat_pocket_depth, carrier_pocket_depth, carrier_slot_depth]
                return valve_pocket
        else:
                pass


    def delete_button(self):
    
        val = tk.IntVar()
        
        stage = [1,2,3,4,5]
        
        option = ttk.OptionMenu(self.frame1, val, *stage)
                              
        option.grid(row = 2, column = 4 , padx =2, pady =2, sticky = "w")
        
        button = ttk. Button(self.frame1, text = "Delete", command = lambda:self.delete_stage(val.get()))
        button.grid(row = 2, column = 3, sticky = "e", padx =2, pady=2)

    
    def delete_stage(self, column):
       
        global input_col_2 , input_col_3 
        print ("delete Column number", column)
        
            
        
    #check whether stage is input_stages and set the column number and delete the column 
              
        if column == 2:
                for item in self.field_stage2:
                        item.destroy()
                for item in self.label_1:
                        item.destroy()
                self.field_stage2 = []
                self.label_1 = []
                input_col_2 = []        
        elif column == 3:
                for item in self.field_stage3:
                        item.destroy()
                for item in self.label_2:
                        item.destroy()        
                self.field_stage3 =[]
                self.label_2 =[]
                input_col_3 =[]
        elif column == 4:
                for item in self.field_stage4:
                        item.destroy()
                for item in self.label_3:
                        item.destroy()          
                self.fied_stage3 =[]
                self.label_3 = []
                input_col_4 = []
        elif column == 5:
                for item in self.field_stage5:
                        item.destroy()
                for item in self.label_4:
                        item.destroy()              
                self.field_stage4 =[]
                self.label_4 =[]
                input_col_5 = []
        #print ("column 2 list", self.field_stage2)
        print ("input column", input_col_2)

   
                
  
""" This is second page for dimension of the valve----------------------------------------------"""
class PageOne(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        self.canvas = tk.Canvas(self, borderwidth=0, background="#efefef")
        self.frame = tk.Frame(self.canvas, background="grey")
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((0,325), window=self.frame, anchor="nw", 
                                  tags="self.frame")
    
      
        self.cage_dim = ["A", "B", "C", "D","E", "Stack Height", "Total Height"]
        self.zvicage_dim = ["Head", "Base", "Column" ,"Cage Height", "Stack Height","Total Height"]
        button1 = ttk.Button(self.frame, text="HOMEPAGE",
                            command=lambda: controller.show_frame(StartPage))
        button1.grid(row = 45, column = 7, sticky ="we")  

            
#-----Cage dimension and part numbers for the cage and column in next page -----------    
class PageTwo(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        
        self.canvas = tk.Canvas(self, borderwidth=0, background="#efefef")
        self.frame = tk.Frame(self.canvas, background="grey")
        self.vsb = tk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.canvas.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side="right", fill="y")
        self.canvas.pack(side="left", fill="both", expand=True)
        self.canvas.create_window((0,325), window=self.frame, anchor="nw", 
                                  tags="self.frame")
    
      
        self.cage_dim = ["A", "B", "C", "D","E", "Stack Height", "Total Height"]
        self.zvicage_dim = ["Head", "Base", "Column" ,"Cage Height", "Stack Height","Total Height"]
        button1 = ttk.Button(self.frame, text="HOMEPAGE",
                            command=lambda: controller.show_frame(StartPage))
        button1.grid(row = 45, column = 7, sticky ="we")  

#helper funcitons to calculate pressure and write data to excel sheet      
   
def press_calc(input_stage):
        #calculates thr pressure difference and pressure ratio
     
         print ("DATA TO WRITE ", input_stage)
         # call function to write data
         
         
         try:   
             amb_press = float(input_stage[3])
             suc_press_abs = float(input_stage[4])+ amb_press
             dis_press_abs = float(input_stage[5])+ amb_press
                
             print (suc_press_abs)
             print (dis_press_abs)
             pressure_diff = round(dis_press_abs - suc_press_abs, 2)
             pressure_ratio = round(dis_press_abs/suc_press_abs, 2)
             return (pressure_diff, pressure_ratio)
         except ValueError:
                return (0, 0)
                


def row_number(row1, sq_number1):
#this function takes the key as the global row number and this will used to write data for all pages 
    global row, sq_number
    row = row1      
    sq_number = sq_number1
    
def compressor_data_excel(comp_data):
# writes the compressor information to excel sheet into base stage excel sheet
    global row
    row1 = row
    print ("ROW TO WRITE DATA", row1)
    
    wb = openpyxl.load_workbook(filename='example1.xlsx', data_only=True)
    ws1 = wb.get_sheet_by_name("BASE STAGE")
    print (comp_data)
    col = 5
    for col, item in enumerate(comp_data, start = 2):
        ws1.cell(row = row1, column = col).value = item
        col += 1
    wb.save("example1.xlsx")
        
    
def base_stage_data1(base_stage, pressure):
#this function writes the frame 3 /operating data of the compressor 
    global  stage 
    print("Base Stage info", base_stage)
    row1 = row
    #find the stage number and open the tab with same stage in workbook
    stage = base_stage[0]
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    ws1.cell(row = row1, column = 1).value = sq_number
    col = 2
    for col, item in enumerate(base_stage, start = 2):
        ws1.cell(row = row1, column = col).value = item
        col += 1    
    col = 10
    for col, item in enumerate(pressure, start =10):
        ws1.cell(row = row1, column = col).value = item
        col += 1
    wb.save("example1.xlsx")
    
def base_stage_data2(seat_design, col):
#writes data to excel sheet for seat design data 
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row
    #check if column is odd or even, if odd write to suction, if even write as discharge 
    
    if col%2 == 1:
            col_excel = 22
    elif col%2 == 0:
            col_excel = 30
            
    for col_excel, item in enumerate(seat_design, start = col_excel):
        ws1.cell(row = row1, column = col_excel).value = item
        col_excel += 1
    wb.save("example1.xlsx")
    
def write_pocket_depth(pocket_depth, col): 
#writes data to excel sheet for pocket depth for both suction and discharge by looking at the column number
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row
    
    if col%2 == 1:
            col_excel = 16
    else:
            col_excel = 19
            
    for col_excel, item in enumerate(pocket_depth, start = col_excel):
        ws1.cell(row = row1, column = col_excel).value = item
        col_excel += 1
    wb.save("example1.xlsx")
    

def write_valve_type(valve_type, col):
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row 
    if col == 1:
        ws1.cell(row = row1, column = 12 ).value = valve_type
       
    elif col == 2:
        ws1.cell(row = row1, column = 14 ).value = valve_type
      
    wb.save("example1.xlsx")    
    
def write_module_data(module_no, col):
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row 
    if col == 1:
        ws1.cell(row = row1, column = 13 ).value = module_no
    elif col == 2:
        ws1.cell(row = row1, column = 15 ).value = module_no
    wb.save("example1.xlsx")
    
def valve_data(col):
    print ("valva data to write", col)
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row 
    if col == 1:
        ws1.cell(row = row1, column = 38 ).value = seat_thick
        ws1.cell(row = row1, column = 39 ).value = carrier_thick
        ws1.cell(row = row1, column = 40 ).value = valve_thick
    elif col == 2:
        ws1.cell(row = row1, column = 41 ).value = seat_thick
        ws1.cell(row = row1, column = 42 ).value = carrier_thick
        ws1.cell(row = row1, column = 43 ).value = valve_thick
    wb.save("example1.xlsx")        

def material_data_excel(fatigue):
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row
    ws1.cell(row = row1, column = 44 ).value = fatigue
    wb.save("example1.xlsx")
    
def stage_data2(seat_design, col):
#writes data to excel sheet for seat design data 
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row
    #check if column is odd or even, if odd write to suction, if even write as discharge 
    
    if col%2 == 1:
            col_excel = 30
    elif col%2 == 0:
            col_excel = 22
            
    for col_excel, item in enumerate(seat_design, start = col_excel):
        ws1.cell(row = row1, column = col_excel).value = item
        col_excel += 1
    wb.save("example1.xlsx")
    
def valve_col_data(seat_thick, carrier_thick, valve_thick,col):
   
    wb = openpyxl.load_workbook(filename='example1.xlsx')
    ws1 = wb.get_sheet_by_name(stage)
    row1 = row 
    print ("column from seat design data", col)
    
    if col% 2 == 0:
        ws1.cell(row = row1, column = 38 ).value = seat_thick
        ws1.cell(row = row1, column = 39 ).value = carrier_thick
        ws1.cell(row = row1, column = 40 ).value = valve_thick
    else:
        ws1.cell(row = row1, column = 41 ).value = seat_thick
        ws1.cell(row = row1, column = 42 ).value = carrier_thick
        ws1.cell(row = row1, column = 43 ).value = valve_thick
    
    wb.save("example1.xlsx")    


# helper functions to write data to UI     
def read_base_data1(row1):
    
    base_data =[]
    print (row1)
    wb = openpyxl.load_workbook(filename='example1.xlsx', data_only=True)
    
    ws = wb.get_sheet_by_name('BASE STAGE')
    for col in range(2,13):
        item = ws.cell(row = row1, column = col).value
        base_data.append(item)
        col  += 1    
    print ("Database data", base_data)
    return base_data
    
def check_stages_data_stage2(row1):
     
     wb = openpyxl.load_workbook(filename='example2.xlsx', data_only=True)
     ws1 = wb.get_sheet_by_name('STAGE 2')
     stage_sq_number1 = ws1.cell(row = row1, column = 1).value
     if sq_number == stage_sq_number1:
         read_stage_data_col2(row1,ws1)
         return True
     else:
         return False
		
def check_stages_data_stage3(row1):		
     wb = openpyxl.load_workbook(filename='example3.xlsx', data_only=True)
     ws1 = wb.get_sheet_by_name('STAGE 3')
     stage_sq_number2 = ws1.cell(row = row1, column = 1).value
     if sq_number ==  stage_sq_number2:
         read_stage_data_col3(row1, ws1)
         return  True
     else:
         return False
        
def read_base_data2(row1):
    
    base_data1 =[]
    print (row1)
    wb = openpyxl.load_workbook(filename='example1.xlsx', data_only=True)
    
    ws = wb.get_sheet_by_name('STAGE 1')
    for col in range(2,44):
        item = ws.cell(row = row1, column = col).value
        base_data1.append(item)
        col  += 1    
    print ("Database data", base_data1)
    return base_data1
    
def read_stage_data_col2(row1, ws):
    global stage_data_col2
    
    for col in range(2,44):
            item = ws.cell(row = row1, column = col).value
            stage_data_col2.append(item)
            col  += 1 
    print ("STAGE 2 DATA", stage_data_col2)
    
    
app = InfoSheet()
app.geometry("1000x800")
app.configure(bg='#334353')
app.title("ZVI-MSA")
app.mainloop()
    

        
